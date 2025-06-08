
from tkinter import messagebox, filedialog
import customtkinter as ctk
import pandas as pd
import openpyxl as px

from Modules.TkSheetTable import TableSheet




class SpelloutARNmbers(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, border_width=1, height=700, width=1100, corner_radius=0)
        self.filepath =''
        self.general_data_frame=''
        self.create_widgets()


    def create_widgets(self):
        self.first_step_frame=ctk.CTkFrame(self, border_width=1, corner_radius=1,height=50)
        self.first_step_frame.pack(fill="x", padx=10, pady=10)

        self.select_file_button = ctk.CTkButton(self.first_step_frame, text="Browse Excel File",font=('Segoe UI Semibold',14), command=self.select_excel_file, width=300, height=30)
        self.select_file_button.pack(side="left", padx=2, pady=2, anchor='w', fill="x")
        self.save_results_btn=ctk.CTkButton(self.first_step_frame, text="Save Results",font=('Segoe UI Semibold',14), command=self.save_results, width=300, height=30)

        self.add_coulmn_spellout=ctk.CTkButton(self.first_step_frame, text="Add Column Spellout",font=('Segoe UI Semibold',14), command=self.add_column_spellout, width=300, height=30)
        
        self.first_step_frame.pack_propagate(False)

        self.data_sheet_frame=ctk.CTkFrame(self, border_width=0.1)

        sheet_name_lable=ctk.CTkLabel(self.data_sheet_frame, text='Select Sheet From File',height=20,corner_radius=1,width=200,font=('Segoe UI Semibold',14),anchor='ne', compound='left',justify='left' )
        sheet_name_lable.grid(row=0, column=0, pady=2, padx=0)

        self.sheet_name = ctk.CTkComboBox(self.data_sheet_frame, values=[], width=200, justify='left',
                             command=self.define_data_frame_and_update_data_sheet)
        self.sheet_name.grid(row=0, column=1, pady=2)

        self.datasheet = TableSheet(self.data_sheet_frame, data=(), width=1500, hight=600, header=[], column_widths=[100], align_center=[0])

     



    # operations functions:___

    def select_excel_file(self):
        self.filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.filepath:
            try:
                self.add_coulmn_spellout.pack(side="left", padx=2, pady=2, anchor='w', fill="x")
                self.data_sheet_frame.pack(fill="x", padx=5, pady=10)
                self.datasheet.grid(row=1, column=0,padx=10, columnspan=5, rowspan=10)
                self.sheet_names = pd.ExcelFile(self.filepath).sheet_names
                self.sheet_name.configure(values=self.sheet_names)
                self.sheet_name.set(self.sheet_names[0])
                self.define_data_frame_and_update_data_sheet(self.sheet_name.get())
            except Exception as e:
                messagebox.showerror('Error', f'حصل الخطأ التالي: {e}')



    def define_data_frame_and_update_data_sheet(self, sheet_name):
        try:
            self.general_data_frame = pd.read_excel(self.filepath, sheet_name=sheet_name)
            self.general_data_frame.dropna(how='all')
            self.general_data_frame.replace(pd.NA,'', inplace=True)
            data=self.general_data_frame.values.tolist()
            self.datasheet.set_header(self.general_data_frame.columns.tolist())
            self.datasheet.load_data(data)
        except Exception as e:
            messagebox.showerror('Error', f'حصل الخطأ التالي: {e}')


    def add_column_spellout(self):
        if  self.general_data_frame.empty:
            messagebox.showerror('خطأ', 'اختر ملف الإكسل أولاً')
            return
        column_list=self.general_data_frame.columns.tolist()
        add_columns_wed = ctk.CTkToplevel()
        add_columns_wed.title("اضبط طريقة التفقيط")
        screen_width = add_columns_wed.winfo_screenwidth()
        screen_height = add_columns_wed.winfo_screenheight()
        window_width = 600
        window_height = 400
        position_x = (screen_width // 2) - (window_width // 2)
        position_y = (screen_height // 2) - (window_height // 2)
        add_columns_wed.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")
        add_columns_wed.focus_set()
        add_columns_wed.resizable(False, True)
        button_frame = ctk.CTkFrame(add_columns_wed)
        button_frame.pack(side='top', padx=10, pady=10, fill='x')
        
        self.numerck_columns_lable=ctk.CTkLabel(button_frame, text='العمود الذي يحتوي على الأرقام', font=('Segoe UI', 14)).pack(side="top", padx=2, pady=2, anchor='w')
        self.columns_selected=ctk.StringVar()
        self.numerck_columns=ctk.CTkComboBox(button_frame,variable=self.columns_selected, values=column_list, width=200).pack(side="top", padx=2, pady=2, anchor='w')

        ctk.CTkLabel(button_frame, text='"ما هو الجمع من المعدود، مثلا:  عشرون "دولاراٌ', font=('Segoe UI', 14)).pack(side="top", padx=2, pady=2, anchor='w')

        self.plural_entry=ctk.CTkEntry(button_frame, font=('Segoe UI', 14), width=200)
        self.plural_entry.pack(side="top", padx=2, pady=2, anchor='w')

        ctk.CTkLabel(button_frame, text='"ما هو المفرد من المعدود، مثلا: خمس "دولارات', font=('Segoe UI', 14)).pack(side="top", padx=2, pady=2, anchor='w')

        self.singular_entry=ctk.CTkEntry(button_frame, font=('Segoe UI', 14), width=200)
        self.singular_entry.pack(side="top", padx=2, pady=2, anchor='w')

        add_btn=ctk.CTkButton(button_frame, text="Add Column",font=('Segoe UI Semibold',14), command=self.add_column, width=300, height=30)
        add_btn.pack(side="top", padx=2, pady=2, anchor='w')


    def add_column(self):
        print(self.columns_selected.get())

        if not self.columns_selected.get():
            messagebox.showerror('خطأ', 'أدخل العمود الذي يحتوي على الأرقام')

        if not self.plural_entry.get() and self.singular_entry.get():
            messagebox.showerror('خطأ', 'أدخل الجمع والمفرد')
            
        else:
            self.general_data_frame['spellout'] = self.general_data_frame[self.columns_selected.get()].apply(
                lambda x: self.arabic_number_to_words(x,  self.singular_entry.get(),self.plural_entry.get()))

            data=self.general_data_frame.values.tolist()
            self.datasheet.set_header(self.general_data_frame.columns.tolist())
            self.datasheet.load_data(data)
            try:
                self.save_results_btn.pack_forget()
            except:pass

            self.save_results_btn.pack(side="left", padx=20, pady=2, anchor='w', fill="x")
            messagebox.showinfo('تم', 'تمت العملية بنجاح')


    def save_results(self):
        wb = px.load_workbook(self.filepath)
        worksheet = wb[self.sheet_name.get()]
        
        last_column = worksheet.max_column + 1
        new_column_name = "spellout"
        worksheet.cell(row=1, column=last_column, value=new_column_name)

        target_column = self.columns_selected.get()

        if isinstance(target_column, int) and target_column > 0:
            pass 
        else:
            for col in range(1, worksheet.max_column +1): 
                if worksheet.cell(row=1, column=col).value == target_column:
                    target_column = col
                    break

        for row in range(2, worksheet.max_row + 1):
            number = worksheet.cell(row=row, column=target_column).value
            converted_value = self.arabic_number_to_words(number,  self.singular_entry.get(), self.plural_entry.get())
            worksheet.cell(row=row, column=last_column, value=converted_value)
        
        wb.save(self.filepath)
        messagebox.showinfo('تم', 'تم حفظ النتائج بنجاح')


    def arabic_number_to_words(self,num,a,b):
        try: int(num)
        except: return 'القيمة في العمود ليست رقما'

        num=int(num)
        full_number=num

        if num == 0:
            return 'صفر'
        below_20 = ['صفر', 'واحد', 'اثنان', 'ثلاثة', 'أربعة', 'خمسة', 'ستة', 'سبعة', 'ثمانية', 'تسعة', 'عشرة', 
                    'أحد عشر', 'اثنا عشر', 'ثلاثة عشر', 'أربعة عشر', 'خمسة عشر', 'ستة عشر', 'سبعة عشر', 
                    'ثمانية عشر', 'تسعة عشر']
        tens = ['', '', 'عشرون', 'ثلاثون', 'أربعون', 'خمسون', 'ستون', 'سبعون', 'ثمانون', 'تسعون']
        hundreds = ['', 'مائة', 'مئتين', 'ثلاثمائة', 'أربعمائة', 'خمسمائة', 'ستمائة', 'سبعمائة', 'ثمانمائة', 'تسعمائة']
        thousands = ['', 'ألف', 'مليون', 'مليار', 'ترليون']
        def helper(n):
            if n == 0:
                return ''
            elif n < 20:
                return below_20[n]
            elif n < 100:
                return (below_20[n % 10] + ' و' if n % 10 != 0 else '') + tens[n // 10] 
           
            elif n < 1000:
                remainder = n % 100
                return hundreds[n // 100] + (' و' + helper(remainder) if remainder != 0 else '')
        def correct_milions(n, unit):
            if n == 1:
                return unit  
            elif n == 2:
                return unit + 'ين'
            elif n >=11:
                return helper(n) + ' ' + unit + ('' if unit != 'ألف' else ' آلاف') 
            else:
                return helper(n) + ' ' + unit + ('ات' if unit != 'ألف' else ' آلاف')  
        def correct_thousands(n, unit):
            if n == 1:
                return unit  
            elif n == 2:
                return unit + 'ين'  
            elif 3 <= n <= 9:
                return below_20[n] + ' آلاف'
            else:
                return helper(n) + ' ألفاً' 
        result = ''
        i = 0
        while num > 0:
            part = num % 1000
            if part != 0:
                if i == 1:
                    part_word = correct_thousands(part, thousands[i])
                elif i >= 2:
                    part_word = correct_milions(part, thousands[i])
                else:
                    part_word = helper(part)
                result = part_word.strip() + ('' if result == '' else ' و' + result)
                #print(f'i: {i}    n-part: {part}     num: {num}')
            num //= 1000
            i += 1
        return result.strip() + f' {a}' if full_number % 100 >=3  and full_number % 100 <=9 else result.strip() + f' {b}'
    


    